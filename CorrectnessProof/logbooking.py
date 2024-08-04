import openpyxl
from openpyxl import Workbook
from datetime import datetime
import os
import time
import main as m
import argparse


def python_print():
    # Load the workbook
    workbook = openpyxl.load_workbook('script_log.xlsx')

    # Select the active sheet
    sheet = workbook.active

    # Print out the values in each row
    for row in sheet.iter_rows(values_only=True):
        print(row)


def clear_workbook():
    filename = 'script_log.xlsx'
    # Load the workbook
    workbook = openpyxl.load_workbook(filename)

    # Iterate through all sheets
    for sheet in workbook.worksheets:
        # Clear the content of each cell in the sheet
        for row in sheet.iter_rows():
            for cell in row:
                cell.value = None

    # Save the workbook
    workbook.save(filename)




def log_to_excel(program_name, iterations, duration, status):
    filename = 'script_log.xlsx'

    # Check if the file exists
    if not os.path.exists(filename):
        # Create a new workbook and add headers
        wb = Workbook()
        ws = wb.active
        ws.title = "Log"
        ws.append(["Program Name", "Iterations", "Duration (s)", "Status", "Message"])
        wb.save(filename)

    # Load the workbook and select the active worksheet
    wb = openpyxl.load_workbook(filename)
    ws = wb.active

    # Find if the program_name already exists
    found = False
    for row in ws.iter_rows(min_row=2, values_only=False):  # Start from the second row to skip headers
        if row[0].value == program_name:
            # Update the row with new values
            row[1].value = iterations
            row[2].value = duration
            row[3].value = status
            row[4].value = "Task completed successfully" if status == "success" else "Task failed"
            found = True
            break

    # If not found, append a new row
    if not found:
        ws.append([program_name, iterations, duration, status,
                   "Task completed successfully" if status == "success" else "Task failed"])

    # Save the workbook
    wb.save(filename)



def stats():
    filename = 'script_log.xlsx'
    # Initialize counters for success and total tasks
    success_count = 0
    total_count = 0

    # Load the workbook
    workbook = openpyxl.load_workbook(filename)
    ws = workbook.active

    # Iterate through all sheets
    for row in ws.iter_rows( values_only=True):
        total_count += 1
        if row[3] == 'success':
            success_count += 1

    # Calculate the success rate
    success_rate = success_count / total_count if total_count > 0 else 0

    print('the success rate is',success_rate)



def main(program_name):

    start_time = time.time()

    try:
        iterations, status = m.main(program_name)



        duration = time.time() - start_time
        log_to_excel(program_name, iterations, duration, status)

        python_print()




    except Exception as e:
        duration = time.time() - start_time
        log_to_excel(program_name, '0', duration, f"Failure: {str(e)}")


def process_file(file_path):
    # Your main processing function for each file
    main(file_path)

def find_c_files(directory):
    c_files = []
    for root, _, files in os.walk(directory):
        for file in files:
            if file.endswith('.c'):
                c_files.append(os.path.join(root, file))
    return c_files

def repeated_file_paths(args):
    for path in args.file_paths:
        if os.path.isfile(path) and path.endswith('.c'):
            process_file(path)
        elif os.path.isdir(path):
            c_files = find_c_files(path)
            for c_file in c_files:
                process_file(c_file)

if __name__ == "__main__":
    #clear_workbook()
    parser = argparse.ArgumentParser()
    parser.add_argument("file_paths", nargs='+', type=str, help="List of file paths or directories to process")
    args = parser.parse_args()
    repeated_file_paths(args)
    stats()


#stats()